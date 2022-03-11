from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
from tkinter.filedialog import askopenfilename
import os

#tkinter frame
root = Tk()
root.title('Portfolio rebalance')
root.geometry("700x350")
file_loc=[]
fol_loc=[]

def choose_dir():
    folder_selected = filedialog.askdirectory()
    Label(root, text= str(folder_selected)).pack()
    fol_loc.append(str(folder_selected))
    return(fol_loc)

def printfn():
    import numpy as np
    import pandas as pd
    import yfinance as yf
    import datetime as dt
    import copy
    import matplotlib.pyplot as plt
    import openpyxl
    import os
    from openpyxl.drawing.image import Image
    
    location = fol_loc[-1]
    file=file_loc[-1]
    filename = str(os.path.basename(file))
    os.chdir(location)
    wb = openpyxl.load_workbook(filename)
    ws_input = wb["Input sheet"]
    wb.remove(wb['Portfolio'])
    wb.create_sheet('Portfolio')
    ws_output = wb["Portfolio"]
    
    num=int(ws_input.cell(row=1, column=4).value)
    rem=int(ws_input.cell(row=2, column=4).value)
    rep=str(ws_input.cell(row=3, column=4).value)
    inter=str(ws_input.cell(row=4, column=4).value)
    bench=str(ws_input.cell(row=5, column=4).value)
    if (inter=="1wk"):
        interval=52
        lab="Weeks"
    if (inter=="1mo"):
        interval=12
        lab="Months"
      
    num_of_tick=0
    for i in range (2,100001):
        if (ws_input.cell(row=i, column=1).value) == None:
             pass
        else:
             num_of_tick+=1
    
    def CAGR(DF):
        df = DF.copy()
        df["cum_return"] = (1 + df["mon_ret"]).cumprod()
        n = len(df)/interval
        CAGR = (df["cum_return"].tolist()[-1])**(1/n) - 1
        return CAGR
    
    def volatility(DF):
        df = DF.copy()
        vol = df["mon_ret"].std() * np.sqrt(interval)
        return vol
    
    def sharpe(DF,rf):
        df = DF.copy()
        sr = (CAGR(df) - rf)/volatility(df)
        return sr
        
    def max_dd(DF):
        df = DF.copy()
        df["cum_return"] = (1 + df["mon_ret"]).cumprod()
        df["cum_roll_max"] = df["cum_return"].cummax()
        df["drawdown"] = df["cum_roll_max"] - df["cum_return"]
        df["drawdown_pct"] = df["drawdown"]/df["cum_roll_max"]
        max_dd = df["drawdown_pct"].max()
        return max_dd
    
    # download historical data
    tickers=[]
    for i in range (2,num_of_tick+2):
        tickers.append(ws_input.cell(row=i, column=1).value)
    
    ohlc_mon = {} # directory with ohlc value for each stock            
    start = dt.datetime.today()-dt.timedelta(3650)
    end = dt.datetime.today()
    for ticker in tickers:
        ohlc_mon[ticker] = yf.download(ticker,start,end,interval=inter)
        ohlc_mon[ticker].dropna(inplace=True,how="all")
     
    tickers = ohlc_mon.keys() # redefine tickers variable after removing any tickers with corrupted data
    
    ################################Backtesting####################################
    
    # calculating ohlc
    ohlc_dict = copy.deepcopy(ohlc_mon)
    return_df = pd.DataFrame()
    for ticker in tickers:
        print("calculating monthly return for ",ticker)
        ohlc_dict[ticker]["mon_ret"] = ohlc_dict[ticker]["Adj Close"].pct_change()
        return_df[ticker] = ohlc_dict[ticker]["mon_ret"]
    
    
    # function to calculate portfolio returns
    portfolio_list=[]
    def pflio(DF,num,rem):
        #DF = dataframe with monthly return info for all stocks
        #num = number of stock in the portfolio
        #rem = number of underperforming stocks to be removed from portfolio
        df = DF.copy()
        portfolio = []
        monthly_ret = [0]
        for i in range(1,len(df)):
            if len(portfolio) > 0:
                monthly_ret.append(df[portfolio].iloc[i,:].mean())
                rem_stocks = df[portfolio].iloc[i,:].sort_values(ascending=True)[:rem].index.values.tolist()
                portfolio = [t for t in portfolio if t not in rem_stocks]
            fill = num - len(portfolio)
            if (rep=="No"):
                new_picks = df[[t for t in tickers if t not in portfolio]].iloc[i,:].sort_values(ascending=False)[:fill].index.values.tolist()
            else:
                new_picks = df.iloc[i,:].sort_values(ascending=False)[:fill].index.values.tolist()
            portfolio = portfolio + new_picks
            portfolio_list.append(portfolio)
        monthly_ret_df = pd.DataFrame(np.array(monthly_ret),columns=["mon_ret"])
        return monthly_ret_df
    
    
    #calculating overall returns
    
    CAGR(pflio(return_df,num,rem))
    sharpe(pflio(return_df,num,rem),0.025)
    max_dd(pflio(return_df,num,rem)) 
    
    #calculating returns for benchmark
    BEN = yf.download(bench,dt.date.today()-dt.timedelta(3650),dt.date.today(),interval=inter)
    BEN["mon_ret"] = BEN["Adj Close"].pct_change()
    CAGR(BEN)
    sharpe(BEN,0.025)
    max_dd(BEN)
    
    #plot
    fig, ax = plt.subplots()
    plt.plot((1+pflio(return_df,num,rem)).cumprod())
    plt.plot((1+BEN["mon_ret"][2:].reset_index(drop=True)).cumprod())
    plt.title("Index Return vs Strategy Return")
    plt.ylabel("Returns x 100%")
    plt.xlabel(lab)
    plt.grid()
    ax.legend(["Strategy Return","Index Return"])
    
    fig.savefig('Backtest - Portfolio rebalance.png')
    img=Image('Backtest - Portfolio rebalance.png')
    img.height=400
    img.width=600
    ws_output.add_image(img,"G2")
    
    
    to_file=portfolio_list[-1]
    # for row in ws_output['A1:C200']:
    #   for cell in row:
    #     cell.value = None
    
    j=2
    ws_output.cell(row=1,column=1).value="Symbol"
    for tof in to_file:
       ws_output.cell(row=j,column=1).value=tof
       j+=1
    
    wb.save(filename)
    wb.close()

    root.destroy()

def open_file():
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '.xlsx .xls')])
    if file:
        filepath = os.path.abspath(file.name)
        Label(root, text= str(filepath)).pack()
        file_loc.append(str(filepath))
        return(file_loc)
      
#add label widget
label = Label(root, text="Click the Button to browse the Files")
label.pack(pady=10)

#buttons
ttk.Button(root, text="Browse directory", command=choose_dir).pack(pady=5)
ttk.Button(root, text="Browse input file", command=open_file).pack(pady=5)
ttk.Button(root, text="Run script", command=printfn).pack(pady=5)

root.mainloop()